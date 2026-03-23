import pandas as pd
from pathlib import Path
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

prompt_techniques = [
    "zeroshot",
    "oneshot",
    "fewshot"
]

models = [
    "Deepseek",
    "Falcon3",
    "Llama",
    "Mistral",
    "Phi3",
    "Qwen",
]

classes = [
    "gender",
    "disability",
    "lgbtq+",
    "race",
    "none"
]

eval_results_path = Path("results") / "subthemes"

def load_classification_report(path: Path):
    if not path.exists():
        return None
    return pd.read_csv(path, index_col=0)


def generate_results_table():

    # colunas (Prompt × Metric)
    columns = pd.MultiIndex.from_product(
        [[p.capitalize() for p in prompt_techniques], ["Pr", "Re", "F1"]],
        names=["Prompt", "Metric"]
    )

    # índice (Class × Model)
    index = pd.MultiIndex.from_product(
        [classes, models],
        names=["Class", "Model"]
    )

    table = pd.DataFrame(index=index, columns=columns)

    for model in models:
        for technique in prompt_techniques:

            tech_cap = technique.capitalize()

            report_path = (
                eval_results_path
                / tech_cap
                / f"{model}_evaluation"
                / "multilabel_classification_report.csv"
            )

            report_df = load_classification_report(report_path)

            if report_df is None:
                continue

            for cls in classes:
                table.loc[(cls, model), (tech_cap, "Pr")] = report_df.loc[cls]["precision"]
                table.loc[(cls, model), (tech_cap, "Re")] = report_df.loc[cls]["recall"]
                table.loc[(cls, model), (tech_cap, "F1")] = report_df.loc[cls]["f1-score"]

    table = table.astype(float).round(2)

    return table

def dataframe_to_excel_styled(df, output_path):

    classes = df.index.get_level_values(0).unique()
    models = df.index.get_level_values(1).unique()
    prompts = ["Zeroshot", "Oneshot", "Fewshot"]
    metrics = ["Pr", "Re", "F1"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Subthemes Results"

    # ===== styles =====
    header_fill = PatternFill("solid", fgColor="9B9B9B")
    zebra_fill = PatternFill("solid", fgColor="C0C0C0")

    header_font = Font(bold=True)
    model_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ===== HEADER 1 =====
    ws.cell(row=1, column=1, value="Class")
    ws.cell(row=1, column=2, value="Model")

    col = 3
    for p in prompts:
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col+2)
        cell = ws.cell(row=1, column=col, value=p)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

        col += 3

    # ===== HEADER 2 =====
    ws.cell(row=2, column=1, value="Class")
    ws.cell(row=2, column=2, value="Model")

    ws.merge_cells(start_row=1, start_column=1,
                       end_row=2, end_column=1)

    c = ws.cell(row=1, column=1, value="Class")
    c.fill = header_fill
    c.font = header_font
    c.alignment = center



    ws.merge_cells(start_row=1, start_column=2,
                       end_row=2, end_column=2)

    c = ws.cell(row=1, column=2, value="Model")
    c.fill = header_fill
    c.font = header_font
    c.alignment = center

    col = 3
    for _ in prompts:
        for m in metrics:
            cell = ws.cell(row=2, column=col, value=m)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            col += 1

    # ===== body =====
    row = 3

    for cls in classes:

        start_class_row = row

        for i, model in enumerate(models):

            ws.cell(row=row, column=2, value=model).font = model_font

            if i % 2 == 1:
                ws.cell(row=row, column=2).fill = zebra_fill

            values = df.loc[(cls, model)]

            col = 3
            for p in prompts:
                for m in metrics:

                    val = values[(p, m)]
                    cell = ws.cell(row=row, column=col, value=f"{float(val):.2f}".replace(",", "."))
                    cell.alignment = center

                    # if i % 2 == 1:
                    #     cell.fill = zebra_fill

                    col += 1

            row += 1

        # merge Class column 
        ws.merge_cells(start_row=start_class_row, start_column=1,
                       end_row=row-1, end_column=1)

        c = ws.cell(row=start_class_row, column=1, value=cls)
        c.font = Font(bold=True)
        c.alignment = center

    # ===== edge =====
    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            ws.cell(row=r, column=c).border = border

    # ===== column width =====
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14

    for i in range(3, max_col+1):
        ws.column_dimensions[get_column_letter(i)].width = 11

    wb.save(output_path)

def save_results_table(output_path):

    df = generate_results_table()

    # df.to_excel(output_path)

    return df


if __name__ == "__main__":

    output_path = Path("results") / "subthemes" / "results_table_subthemes.xlsx"

    table = save_results_table(output_path)

    latex_code = dataframe_to_excel_styled(
        table,
        output_path
    )

    print(table)